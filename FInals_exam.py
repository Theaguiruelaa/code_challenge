from tkinter import *
import tkinter as tk
from openpyxl import Workbook, load_workbook
import os

window = tk.Tk()
window.title("Grades")
window.geometry("300x300")

label1 = tk.Label(window, text="Grades", font=("Arial", 20))
label1.pack()

label3 = tk.Label(window, text="Name:")
label3.pack()

user = tk.Entry(window, textvariable="Name", width=30)
user.pack()

label2 = tk.Label(window, text="Course:")
label2.pack()

user2 = tk.Entry(window, textvariable="Course", width=30)
user2.pack()

label5 = tk.Label(window, text="Grade:")
label5.pack()

user5 = tk.Entry(window, textvariable="Grade", width=30)
user5.pack()

filename = "Grades.xlsx"

def create_excel():
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"
        ws.append(["Name", "Course", "Grade"])

        wb.save(filename)

def save_data():
    name = user.get()
    course = user2.get()
    grade = user5.get()

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        row_values = [name, course, grade]
        ws.append(row_values)
        wb.save(filename)


but1 = tk.Button(window, text="Save", command=save_data)
but1.pack()

create_excel()
window.mainloop()