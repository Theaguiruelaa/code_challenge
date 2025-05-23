from tkinter import *
import tkinter as tk
import tkinter.ttk as ttk
from openpyxl import Workbook, load_workbook
import os

window = tk.Tk()
window.title("Student Grades Manager")
window.geometry("400x350")
window.resizable(False, False)

label1 = tk.Label(window, text="Student Grades Manager", font=("Arial", 16))
label1.pack(pady=10)

label3 = tk.Label(window, text="Name:")
label3.pack()
user = tk.Entry(window, width=40)
user.pack()

label2 = tk.Label(window, text="Course:")
label2.pack()
user2 = tk.Entry(window, width=40)
user2.pack()

label5 = tk.Label(window, text="Grade:")
label5.pack()

def validate_grade(value):
    if value == "":
        return True
    try:
        float(value)
        return True
    except ValueError:
        return False

vcmd = window.register(validate_grade)
user5 = tk.Entry(window, width=40, validate="key", validatecommand=(vcmd, "%P"))
user5.pack()

filename = "Grades.xlsx"

def create_excel():
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"
        ws.append(["Name", "Course", "Grade"])
        wb.save(filename)

def save_data():
    name = user.get()
    course = user2.get()
    grade = user5.get()

    if name and course and grade:
        try:
            float(grade)
        except ValueError:
            return

        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
            ws.append([name, course, grade])
            wb.save(filename)

def view_grades():
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active

        view_win = Toplevel(window)
        view_win.title("View Grades")
        view_win.geometry("400x300")
        view_win.resizable(False, False)

        style = ttk.Style(view_win)
        style.theme_use('default')

        style.configure("Treeview.Heading", background="#1560BD", foreground="white", font=("Arial", 10, "bold"))
        style.map('Treeview', background=[('selected', '#347083')])

        tree = ttk.Treeview(view_win, columns=("Name", "Course", "Grade"), show="headings", height=12)
        tree.pack(fill="both", expand=True)

        tree.heading("Name", text="Name")
        tree.heading("Course", text="Course")
        tree.heading("Grade", text="Grade")

        tree.column("Name", width=140, anchor="w")
        tree.column("Course", width=140, anchor="w")
        tree.column("Grade", width=80, anchor="center")

        tree.tag_configure('oddrow', background='#cce6ff')
        tree.tag_configure('evenrow', background='#99ccff')

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is not None for cell in row):
                tag = "evenrow" if count % 2 == 0 else "oddrow"
                tree.insert("", "end", values=row, tags=(tag,))
                count += 1

but1 = tk.Button(window, text="Save", command=save_data, bg="#86B049")
but1.pack(pady=5)

but2 = tk.Button(window, text="View Grades", command=view_grades, bg="#1560BD")
but2.pack(pady=5)

create_excel()
window.mainloop()